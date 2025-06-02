
from flask import Flask, render_template, request, redirect, url_for, flash
import os
from openpyxl import load_workbook
from collections import defaultdict
from shipdate_helper import set_ship_date, get_all_files_with_ship_dates


app = Flask(__name__)
app.secret_key = 'your_secret_key'
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route('/')
def upload_dashboard():
    files = get_all_files_with_ship_dates()
    return render_template('upload_dashboard.html', files=files)

@app.route('/upload', methods=['POST'])
def upload():
    file = request.files['file']
    ship_date = request.form.get('ship_date')
    if file:
        path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(path)
        set_ship_date(file.filename, ship_date)
        return redirect(url_for('scan_items', filename=file.filename))
    return redirect(url_for('upload_dashboard'))

@app.route('/packing_list/<filename>', methods=['GET', 'POST'])
def scan_items(filename):
    path = os.path.join(UPLOAD_FOLDER, filename)
    wb = load_workbook(path)
    sheet = wb.active

    barcode = request.form.get('barcode')
    if barcode:
        found = False
        for row in sheet.iter_rows(min_row=2):
            if str(row[0].value).strip() == barcode:
                found = True
                scanned_qty = row[3].value or 0
                row[3].value = scanned_qty + 1
                if not row[6].value:
                    row[6].value = get_next_pallet_number(sheet)
                break
        if not found:
            flash(f"⚠️ Barcode '{barcode}' not found in Master List.")
        wb.save(path)

    if 'new_pallet' in request.form:
        if any(row[6].value is None and row[3].value for row in sheet.iter_rows(min_row=2)):
            flash("⚠️ Cannot start a new pallet. You have unassigned scanned items.")
        else:
            next_pallet = get_next_pallet_number(sheet)
            flash(f"Pallet {next_pallet} started. You can now begin scanning.")

    master_list_rows = []
    pallet_rows = defaultdict(list)
    for row in sheet.iter_rows(min_row=2):
        part = row[0].value
        desc = row[1].value
        order = row[2].value
        scanned_qty = row[3].value or 0
        ship_qty = row[4].value or 0
        order_qty = row[5].value or 0
        pallet = row[6].value
        bo_qty = order_qty - ship_qty

        values = (part, desc, order, scanned_qty, ship_qty, order_qty, pallet, bo_qty)
        master_list_rows.append(values)
        if pallet:
            pallet_rows[str(pallet)].append(values)

    grouped_rows = {"Master List": master_list_rows}
    grouped_rows.update(pallet_rows)
    can_start = any(row[6].value for row in sheet.iter_rows(min_row=2))

    return render_template('packing_list.html', filename=filename,
                           grouped_rows=grouped_rows,
                           can_start_new_pallet=can_start)

def get_next_pallet_number(sheet):
    pallets = {row[6].value for row in sheet.iter_rows(min_row=2) if row[6].value}
    nums = [int(p) for p in pallets if str(p).isdigit()]
    return max(nums, default=0) + 1

@app.route('/reset/<filename>', methods=['POST'])
def reset_scanned(filename):
    path = os.path.join(UPLOAD_FOLDER, filename)
    wb = load_workbook(path)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2):
        row[3].value = 0
        row[6].value = None
    wb.save(path)
    flash("All scanned quantities reset.")
    return redirect(url_for('scan_items', filename=filename))

@app.route('/delete_pallet/<filename>/<pallet_name>', methods=['POST'])
def delete_pallet(filename, pallet_name):
    path = os.path.join(UPLOAD_FOLDER, filename)
    wb = load_workbook(path)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2):
        if str(row[6].value) == str(pallet_name):
            row[3].value = 0
            row[6].value = None
    wb.save(path)
    flash(f"Pallet {pallet_name} deleted.")
    return redirect(url_for('scan_items', filename=filename))

@app.route('/delete_row/<filename>', methods=['POST'])
def delete_row(filename):
    barcode = request.form['barcode']
    pallet = request.form['pallet']
    path = os.path.join(UPLOAD_FOLDER, filename)
    wb = load_workbook(path)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2):
        if str(row[0].value) == barcode and str(row[6].value) == pallet:
            row[3].value = 0
            row[6].value = None
    wb.save(path)
    flash(f"Removed {barcode} from pallet {pallet}")
    return redirect(url_for('scan_items', filename=filename))

if __name__ == '__main__':
    app.run(debug=True)
